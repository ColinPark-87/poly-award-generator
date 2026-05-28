from __future__ import annotations

import csv
import re
import unicodedata
import openpyxl
from typing import Any

MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]

COL_CLASS         = 1
COL_LEVEL         = 2
COL_NAME          = 4
COL_ENGLISH       = 6   # English
COL_SPEECH        = 7   # Speech Building
COL_FOUNDATIONS   = 8   # Eng. Foundations
COL_LC            = 9   # Lang. Composition
COL_NF            = 10  # NF Studies
COL_TOTAL         = 11  # TOTAL (English+Speech+Foundations 합계, LC/NF 미포함)
COL_AVERAGE       = 12
COL_CLASS_RANKING = 13
COL_LEVEL_RANKING = 14

# 정발 캠퍼스 점수 가중치 키 — UI/config/계산 모두에서 같은 키 사용
JUNGBAL_SUBJECT_KEYS = ("english", "speech", "foundations", "lc", "nf")
JUNGBAL_DEFAULT_WEIGHTS: dict[str, float] = {
    "english":     1.0,
    "speech":      1.0,
    "foundations": 1.0,
    "lc":          1.0,
    "nf":          0.0,
}


def clean_class_name(cls: str) -> str:
    """클래스명에서 요일/시간 표기 제거.
    예: 'GT3 (MWF)3:10'   → 'GT3'
        'MGT2(t/th)'       → 'MGT2'
        'S1 (M/W/F) 4:20'  → 'S1'
        'GT2 4:40'         → 'GT2'
        'GT1-보스턴（화목）' → 'GT1-Boston' (NFKC 전각→ASCII, 한국어 제거)
    """
    # NFKC: 전각 문자 → ASCII 등가 (Ａ→A, （→(, ）→) 등)
    s = unicodedata.normalize('NFKC', str(cls).strip())
    s = re.sub(r'\s*\d+:\d+', '', s)     # 시간 표기 먼저 제거 (괄호보다 앞에: "(MWF)3:10"에서 숫자 오버랩 방지)
    s = re.sub(r'\s*\([^)]*\)', '', s)   # ASCII 괄호 그룹 제거
    # 라틴/숫자/기본 구두점 외 문자(한국어 등) 제거
    s = re.sub(r'[^\x00-\x7FÀ-ɏ]', '', s).strip()
    return s if s else str(cls).strip()


def parse_student_name(full_name: str) -> tuple[str, str]:
    """'최윤아 (Elena Choi)' → ('최윤아', 'Elena Choi')"""
    # NFKC: 전각 문자 정규화 후 괄호 추출 (Ｅlena → Elena 등)
    s = unicodedata.normalize('NFKC', str(full_name))
    match = re.search(r'\(([^)]+)\)', s)
    if match:
        korean = s[:match.start()].strip()
        english = match.group(1).strip()
        return korean, english
    return s.strip(), s.strip()


def extract_month_from_filename(filename: str) -> str:
    """파일명에서 'April 2026' 형태 추출"""
    pattern = r'(' + '|'.join(MONTHS) + r')\s+(\d{4})'
    match = re.search(pattern, filename)
    if match:
        return f"{match.group(1)} {match.group(2)}"
    return ""


def _to_int(v: Any) -> int:
    """엑셀 셀 → int. None/빈 문자열/숫자 아닌 값은 0."""
    if v is None:
        return 0
    try:
        return int(v)
    except (TypeError, ValueError):
        return 0


def load_rows_from_excel(file_path: str) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # 헤더(2행)에서 열 위치 자동 감지 — MT/LT 파일 구조 차이 대응
    # MT: TOTAL=col11, AVG=col12, CLASS_RANK=col13, LEVEL_RANK=col14, LC=col9
    # LT: TOTAL=col19, AVG=col20, CLASS_RANK=col23, LEVEL_RANK=col24, LC=col17
    header = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))

    def _find_col(name: str, default: int) -> int:
        for i, h in enumerate(header):
            if h and str(h).strip() == name:
                return i
        return default

    col_lc       = _find_col("Lang. Composition", COL_LC)
    col_total    = _find_col("TOTAL",         COL_TOTAL)
    col_avg      = _find_col("Average",       COL_AVERAGE)
    col_cls_rank = _find_col("Class Ranking", COL_CLASS_RANKING)
    col_lvl_rank = _find_col("Level Ranking", COL_LEVEL_RANKING)
    col_english  = _find_col("English",       COL_ENGLISH)
    col_speech   = _find_col("Speech Building", COL_SPEECH)
    col_fnd      = _find_col("Eng. Foundations", COL_FOUNDATIONS)
    col_nf       = _find_col("NF Studies",    COL_NF)
    col_sr       = _find_col("StarReading",   -1)   # LT 파일에만 존재; MT 파일은 -1

    def _to_float_safe(v: Any) -> float:
        if v is None:
            return 0.0
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        name_raw  = row[COL_NAME]
        lc_raw    = row[col_lc]       if len(row) > col_lc       else None
        total_raw = row[col_total]    if len(row) > col_total    else None
        avg_raw   = row[col_avg]      if len(row) > col_avg      else None
        cls_raw   = row[COL_CLASS]
        level_raw = row[COL_LEVEL]    if len(row) > COL_LEVEL    else None
        cls_rank  = row[col_cls_rank] if len(row) > col_cls_rank else None
        lvl_rank  = row[col_lvl_rank] if len(row) > col_lvl_rank else None
        eng_raw   = row[col_english]  if len(row) > col_english  else None
        sp_raw    = row[col_speech]   if len(row) > col_speech   else None
        fnd_raw   = row[col_fnd]      if len(row) > col_fnd      else None
        nf_raw    = row[col_nf]       if len(row) > col_nf       else None
        sr_raw    = row[col_sr]       if col_sr >= 0 and len(row) > col_sr else None
        if not name_raw or avg_raw is None:
            continue
        try:
            lc    = _to_int(lc_raw)
            total = _to_int(total_raw)
            avg   = float(avg_raw)
        except (TypeError, ValueError):
            continue
        rows.append({
            "class":         clean_class_name(str(cls_raw).strip()) if cls_raw else "",
            "level":         str(level_raw).strip() if level_raw else "",
            "name":          str(name_raw).strip(),
            "english":       _to_int(eng_raw),
            "speech":        _to_int(sp_raw),
            "foundations":   _to_int(fnd_raw),
            "lc":            lc,
            "nf":            _to_int(nf_raw),
            "sr":            _to_float_safe(sr_raw),   # LT: GE값, MT: 0.0
            "total":         total,
            "average":       avg,
            "class_ranking": str(cls_rank).strip() if cls_rank else "",
            "level_ranking": str(lvl_rank).strip() if lvl_rank else "",
        })
    return rows


def _parse_ge(ge_str) -> float | None:
    """GE 값 파싱. '-' 또는 비어있으면 None, '>12.9' 형태도 처리."""
    s = str(ge_str).strip() if ge_str else ""
    if s in ("-", ""):
        return None
    s = s.lstrip(">")
    try:
        return float(s)
    except ValueError:
        return None


def load_sr_from_csv(file_path: str) -> list[dict[str, Any]]:
    """
    Star Summary Report CSV 파싱.
    - Student 컬럼: '학번, 이름' → 이름만 추출
    - GE '-' 또는 test 반은 제외
    """
    rows = []
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            cls        = row.get("Class/Group", "").strip()
            student_raw = row.get("Student", "").strip()
            ge_raw     = row.get("GE", "-")

            if not cls or not student_raw:
                continue
            if cls.lower().startswith("test"):
                continue

            ge = _parse_ge(ge_raw)
            if ge is None:          # 점수 없는 학생·반 제외
                continue

            # '학번, 이름' → 이름만
            parts = student_raw.split(",", 1)
            name  = parts[1].strip() if len(parts) == 2 else student_raw

            rows.append({
                "class":        cls,
                "english_name": name,
                "ge":           ge,
            })
    return rows


def load_sr_from_excel_yuseong(file_path: str) -> list[dict[str, Any]]:
    """
    유성 SR Excel 파싱.
    - Sheet "SR", 헤더 2행, 데이터 3행~
    - Class 셀 병합 → carry-forward
    - 이름: "한국이름 (English Name)" → 영문만 추출, 후행 # 제거
    - Class: "ClassName\\nTeacher" 또는 "ClassName   Teacher" → 앞부분만
    - 최고 GE: None/'<K' → 월별 최대값 fallback, 그래도 없으면 학생 제외
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb["SR"] if "SR" in wb.sheetnames else wb.active

    header = [str(c.value).strip() if c.value is not None else "" for c in ws[2]]

    def _col(name: str, default: int = -1) -> int:
        for i, h in enumerate(header):
            if h == name:
                return i
        return default

    col_name    = _col("Name", 1)
    col_class   = _col("Class", 2)
    col_best_ge = _col("최고 GE", 3)
    month_cols  = [i for i, h in enumerate(header) if h in MONTHS]

    def _parse_class_raw(raw) -> str:
        s = str(raw).strip()
        if "\n" in s:
            return s.split("\n")[0].strip()
        return re.split(r"\s{2,}", s)[0].strip()

    def _parse_ge_val(v) -> float | None:
        if v is None:
            return None
        s = str(v).strip()
        if s in ("-", "", "<K"):
            return None
        s = s.lstrip(">")
        try:
            return float(s)
        except ValueError:
            return None

    def _english_name(raw: str) -> str:
        s = unicodedata.normalize('NFKC', str(raw).strip().lstrip("\n").strip())
        m = re.search(r'\(([^)]+)\)', s)
        en = m.group(1).strip() if m else s
        return en.rstrip("#* \t").strip()

    rows: list[dict[str, Any]] = []
    current_class = ""

    for row in ws.iter_rows(min_row=3, values_only=True):
        if col_name >= len(row):
            continue
        name_raw = row[col_name]
        if not name_raw:
            continue
        name_str = str(name_raw)
        # 중복 헤더 행 건너뜀
        if name_str.strip() in ("Name", "Student"):
            continue

        # Class carry-forward
        cls_val = row[col_class] if col_class < len(row) else None
        if cls_val:
            current_class = _parse_class_raw(cls_val)
        if not current_class:
            continue

        # GE 계산
        ge = _parse_ge_val(row[col_best_ge]) if col_best_ge < len(row) else None
        if ge is None and month_cols:
            monthly = [_parse_ge_val(row[c]) for c in month_cols if c < len(row)]
            valid   = [v for v in monthly if v is not None]
            ge = max(valid) if valid else None
        if ge is None:
            continue

        rows.append({
            "class":        current_class,
            "english_name": _english_name(name_str),
            "ge":           ge,
        })

    return rows


def select_sr_winners(rows: list[dict[str, Any]]) -> list[dict]:
    """Class/Group별 GE 최고점 학생 1명 선정. 점수 없는 반은 자동 제외."""
    winner_map: dict[str, dict] = {}
    for row in rows:
        cls = row["class"]
        if cls not in winner_map or row["ge"] > winner_map[cls]["ge"]:
            winner_map[cls] = row
    return list(winner_map.values())


def _get_level(cls: str) -> str:
    """클래스명에서 레벨 접두어 추출. 예: 'GT3' → 'GT', 'MGT2' → 'MGT'"""
    m = re.match(r"^(GT|MGT|MAG|S)", cls)
    return m.group(1) if m else ""


# ── 분당엠폴리 ─────────────────────────────────────────────
def _bd_norm(v: Any) -> str:
    return (str(v).replace("\n", " ").strip() if v is not None else "")


def load_bundang_level_top(file_path: str) -> list[dict[str, Any]]:
    """초등TOP + 중등TOP 시트에서 Level TOP == 1 학생 → Level Top 상장 명단.
    시트는 레벨 그룹마다 헤더가 반복되는 블록 구조."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    out: list[dict[str, Any]] = []
    for sn in ("초등TOP", "중등TOP"):
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        colmap: dict[str, int] | None = None
        for row in ws.iter_rows(values_only=True):
            cells = [_bd_norm(c) for c in row]
            if "학급명" in cells:
                colmap = {name: i for i, name in enumerate(cells) if name}
                continue
            if not colmap:
                continue
            ci = colmap.get("학급명")
            if ci is None or ci >= len(cells) or not cells[ci]:
                continue

            def _g(key: str) -> str:
                i = colmap.get(key)
                return cells[i] if i is not None and i < len(cells) else ""

            if _g("Level TOP") == "1":
                kor, eng = parse_student_name(_g("학생이름"))
                out.append({
                    "class":        _g("학급명").strip(),
                    "korean_name":  kor,
                    "english_name": eng,
                    "full_name":    _g("학생이름").strip(),
                })
    return out


def load_bundang_grammar(file_path: str) -> list[dict[str, Any]]:
    """종합성적관리 시트에서 Eng. Mechanics 만점(컬럼 최댓값) 학생 → Grammar 상장 명단."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if "종합성적관리" not in wb.sheetnames:
        return []
    ws = wb["종합성적관리"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    hdr = [_bd_norm(c) for c in rows[0]]
    colmap = {name: i for i, name in enumerate(hdr) if name}
    ci = colmap.get("학급", colmap.get("학급명"))
    ni = colmap.get("학생이름")
    ei = colmap.get("Eng. Mechanics")
    if ni is None or ei is None:
        return []

    recs: list[tuple[str, str, float]] = []
    for row in rows[1:]:
        cells = [_bd_norm(c) for c in row]
        if ei >= len(cells) or not cells[ei]:
            continue
        try:
            em = float(cells[ei])
        except ValueError:
            continue
        cls  = cells[ci] if ci is not None and ci < len(cells) else ""
        name = cells[ni] if ni < len(cells) else ""
        if name:
            recs.append((cls.strip(), name.strip(), em))
    if not recs:
        return []

    max_em = max(r[2] for r in recs)
    out: list[dict[str, Any]] = []
    for cls, full_name, em in recs:
        if em == max_em:
            kor, eng = parse_student_name(full_name)
            out.append({
                "class":        cls,
                "korean_name":  kor,
                "english_name": eng,
                "full_name":    full_name,
            })
    return out


def select_winners(
    rows: list[dict[str, Any]],
    perfect_score_min: float = 100.0,
    honor_roll_min:    float = 95.0,
    best_writer_min_lc = 0,
) -> dict[str, list[dict]]:
    """
    수상자 선정.
    perfect_score_min  : 이 값 이상이면 Perfect Score (기본 100)
    honor_roll_min     : 이 값 이상 ~ perfect_score_min 미만이면 Honor Roll (기본 95)
    best_writer_min_lc : Best Writer 자격 최소 LC 점수.
                         int → 전 레벨 공통 기준
                         dict → {'GT': 27, 'MGT': 25, 'S': 20, 'MAG': 20} 형태로 레벨별 기준
    """
    perfect_score = []
    honor_roll    = []
    best_writer_map: dict[str, dict] = {}

    for row in rows:
        avg = row["average"]
        lc  = row["lc"]
        cls = row["class"]
        korean, english = parse_student_name(row["name"])
        student = {
            "korean_name":  korean,
            "english_name": english,
            "class":        cls,
            "average":      avg,
            "lc":           lc,
        }

        if avg >= perfect_score_min:
            perfect_score.append(student)
        elif avg >= honor_roll_min:
            honor_roll.append(student)

        # Best Writer: 학급별 LC 최고점, 동점시 Average 높은 1명
        if isinstance(best_writer_min_lc, dict):
            level  = _get_level(cls)
            min_lc = best_writer_min_lc.get(level, 0)
        else:
            min_lc = best_writer_min_lc

        if lc >= min_lc:
            if cls not in best_writer_map:
                best_writer_map[cls] = student
            else:
                cur = best_writer_map[cls]
                if lc > cur["lc"] or (lc == cur["lc"] and avg > cur["average"]):
                    best_writer_map[cls] = student

    return {
        "perfect_score": perfect_score,
        "honor_roll":    honor_roll,
        "best_writer":   list(best_writer_map.values()),
    }


def select_jungbal_winners(
    rows: list[dict[str, Any]],
    weights: dict[str, float] | None = None,
) -> dict[str, list[dict]]:
    """
    정발 캠퍼스 수상자 선정.

    weights: 과목별 가중치 dict.
        {"english": 1.0, "speech": 1.0, "foundations": 1.0, "lc": 1.0, "nf": 0.0}
        None이면 JUNGBAL_DEFAULT_WEIGHTS 사용.
        지정되지 않은 과목은 0으로 취급.

    로직 (모든 과목을 가중치로 합산, 동점은 공동 수상):
    1) 반별로 가중 합산 점수가 가장 높은 학생 = 반 1등 (동점이면 공동).
       (Excel Class Ranking 열은 LC/NF 미반영이라 사용하지 않음)
    2) 반 1등을 level(Col 2)별로 그룹화
    3) 레벨 내 반 1등 중 합산 점수 최고와 같은 학생 → Achievement Certificate (공동 가능)
       나머지 반 1등 → Monthly Test Winner
    """
    from collections import defaultdict

    w = dict(JUNGBAL_DEFAULT_WEIGHTS)
    if weights:
        for k, v in weights.items():
            try:
                w[k] = float(v)
            except (TypeError, ValueError):
                pass

    def _score(r: dict) -> float:
        # total = TOTAL 열(LC 제외 전 과목 합계), lc = Lang. Composition
        # 개별 과목 가중치 대신 TOTAL+LC 직접 사용 — MT/LT 파일 모두 정확
        return r.get("total", 0) + r.get("lc", 0)

    def _sr(r: dict) -> float:
        return r.get("sr", 0.0)

    # ── Step 1: 반별 최고점 학생 수집 ────────────────────────────────────
    # 우선순위: Total+LC 높은 순 → SR 높은 순 → SR도 동점이면 공동 수상
    by_class: dict[str, list[dict]] = {}
    for row in rows:
        cls = row["class"]
        if cls not in by_class:
            by_class[cls] = [row]
        else:
            best_score = _score(by_class[cls][0])
            best_sr    = _sr(by_class[cls][0])
            row_score  = _score(row)
            row_sr     = _sr(row)
            if row_score > best_score:
                by_class[cls] = [row]
            elif row_score == best_score:
                if row_sr > best_sr:
                    by_class[cls] = [row]       # SR 높아 단독 승
                elif row_sr == best_sr:
                    by_class[cls].append(row)   # SR도 동점 → 공동

    # ── Step 2: 레벨별 그룹화 ──────────────────────────────────────────
    by_level: dict[str, dict[str, list[dict]]] = defaultdict(dict)
    for cls, cls_rows in by_class.items():
        by_level[cls_rows[0]["level"]][cls] = cls_rows

    # ── Step 3: Achievement vs Monthly 분류 ─────────────────────────────
    achievement: list[dict] = []
    monthly_winner: list[dict] = []

    def _make_student(row: dict) -> dict:
        korean, english = parse_student_name(row["name"])
        return {
            "korean_name":   korean,
            "english_name":  english,
            "class":         row["class"],
            "level":         row["level"],
            "average":       row["average"],
            "lc":            row["lc"],
            "total":         row["total"],
            "score":         _score(row),
            "class_ranking": row["class_ranking"],
            "level_ranking": row["level_ranking"],
        }

    for level, classes in by_level.items():
        if len(classes) == 1:
            # 단일 반 레벨 → Achievement (동점자 전원)
            for row in next(iter(classes.values())):
                achievement.append(_make_student(row))
            continue
        # 복수 반: 레벨 내 최고 점수 반 → Achievement, 나머지 → Monthly
        # 반 대표 점수 = (총점, 반 내 최고 SR) — 레벨 간 동점은 SR로 해소
        best_cls = max(
            classes,
            key=lambda c: (_score(classes[c][0]), max(_sr(r) for r in classes[c])),
        )
        for cls, cls_rows in classes.items():
            for row in cls_rows:
                s = _make_student(row)
                if cls == best_cls:
                    achievement.append(s)
                else:
                    monthly_winner.append(s)

    return {
        "achievement_certificate": achievement,
        "monthly_test_winner":     monthly_winner,
    }
